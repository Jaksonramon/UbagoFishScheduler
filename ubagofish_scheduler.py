"""
Ubagofish Scheduler — Versioned Script
Version: 2.0
Changelog:
- v1: Core scheduler with balanced-day randomizer, configurable work-rest cadence, manual-locked appointments, and Excel export.
- v2.0: Reinforced "locked" flag for manual appointments (never overwritten), improved save/load JSON, per-day Excel export with ByBuyer/ByClient sheets and lunch slots greyed out.

Notes:
- Manual (locked) appointments are preserved and never overwritten by the randomizer.
- Randomizer only removes/rewires unlocked appointments for selected buyers.
- Export generates one pair of sheets per selected day: `ByBuyer_{day}` and `ByClient_{day}`.
"""

import streamlit as st
import pandas as pd
import json
import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# -------------------------
# App config & constants
# -------------------------
st.set_page_config(page_title="UbagoFish Scheduler v2.0", layout="wide")

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
HOURS = [f"{h:02d}:{m:02d}" for h in range(6, 22) for m in (0, 30)]
DATA_FILE = "ubagofish_data.json"

# -------------------------
# Session defaults
# -------------------------
for key in ["clients", "buyers", "appointments"]:
    if key not in st.session_state:
        st.session_state[key] = []  # appointments: list of dicts {client,buyer,day,time,locked}
if "edit_expander_open" not in st.session_state:
    st.session_state.edit_expander_open = False
if "start_hour" not in st.session_state:
    st.session_state.start_hour = "08:00"
if "end_hour" not in st.session_state:
    st.session_state.end_hour = "18:00"
if "lunch_start" not in st.session_state:
    st.session_state.lunch_start = "12:00"
if "lunch_end" not in st.session_state:
    st.session_state.lunch_end = "14:00"
if "selected_days" not in st.session_state:
    st.session_state.selected_days = ["Monday", "Tuesday"]
if "time_windows" not in st.session_state:
    st.session_state.time_windows = {}  # {buyer: {day: {start,end}}}

# -------------------------
# Persistence helpers
# -------------------------

def load_data_from_disk():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return
        st.session_state.clients = data.get("clients", st.session_state.clients)
        st.session_state.buyers = data.get("buyers", st.session_state.buyers)
        # normalize appointments and locked flag
        appts_in = data.get("appointments", [])
        appts = []
        for a in appts_in:
            if isinstance(a, dict):
                locked = a.get("locked")
                if locked is None and "manual" in a:
                    locked = bool(a.get("manual", False))
                appts.append({
                    "client": a.get("client") or a.get("Client"),
                    "buyer": a.get("buyer") or a.get("Buyer"),
                    "day": a.get("day") or a.get("Día") or a.get("day"),
                    "time": a.get("time") or a.get("Hora"),
                    "locked": bool(locked)
                })
            else:
                try:
                    client, buyer, day, time = a
                    appts.append({"client": client, "buyer": buyer, "day": day, "time": time, "locked": False})
                except Exception:
                    continue
        st.session_state.appointments = appts
        st.session_state.start_hour = data.get("start_hour", st.session_state.start_hour)
        st.session_state.end_hour = data.get("end_hour", st.session_state.end_hour)
        st.session_state.lunch_start = data.get("lunch_start", st.session_state.lunch_start)
        st.session_state.lunch_end = data.get("lunch_end", st.session_state.lunch_end)
        st.session_state.selected_days = data.get("selected_days", st.session_state.selected_days)
        st.session_state.time_windows = data.get("time_windows", st.session_state.time_windows)


def save_data_to_disk():
    data = {
        "clients": st.session_state.clients,
        "buyers": st.session_state.buyers,
        "appointments": st.session_state.appointments,
        "start_hour": st.session_state.start_hour,
        "end_hour": st.session_state.end_hour,
        "lunch_start": st.session_state.lunch_start,
        "lunch_end": st.session_state.lunch_end,
        "selected_days": st.session_state.selected_days,
        "time_windows": st.session_state.time_windows,
    }
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def autosave():
    try:
        save_data_to_disk()
    except Exception:
        pass

# initial load
load_data_from_disk()

# -------------------------
# Time helpers & constraints
# -------------------------

def idx_of(t: str) -> int:
    return HOURS.index(t)

lunch_start_idx, lunch_end_idx = idx_of(st.session_state.lunch_start), idx_of(st.session_state.lunch_end)

def is_in_lunch_break(t: str) -> bool:
    return lunch_start_idx <= HOURS.index(t) < lunch_end_idx

# remove any appointment accidentally saved during lunch
st.session_state.appointments = [a for a in st.session_state.appointments if not is_in_lunch_break(a["time"])]


def is_slot_free(client: str, buyer: str, day: str, time: str) -> bool:
    """Slot is free if no appointment exists for same day/time with either buyer or client."""
    for a in st.session_state.appointments:
        if a["day"] == day and a["time"] == time:
            if a["buyer"] == buyer or a["client"] == client:
                return False
    return True

# -------------------------
# Sidebar: config + save/load
# -------------------------
with st.sidebar:
    st.header("Configuration & Data")
    buyers_input = st.text_area("Buyers (uno por línea)", "
".join(st.session_state.buyers), height=180)
    st.session_state.buyers = [b.strip() for b in buyers_input.splitlines() if b.strip()]
    clients_input = st.text_area("Clients (uno por línea)", "
".join(st.session_state.clients), height=180)
    st.session_state.clients = [c.strip() for c in clients_input.splitlines() if c.strip()]

    if st.button("Guardar nombres"):
        autosave(); st.success("Datos guardados en sesión y disco.")

    st.subheader("Días y Horario")
    st.session_state.selected_days = st.multiselect("Días a programar", DAYS, default=st.session_state.selected_days)
    st.session_state.start_hour = st.selectbox("Inicio del día", HOURS, index=idx_of(st.session_state.start_hour))
    st.session_state.end_hour = st.selectbox("Fin del día", HOURS, index=idx_of(st.session_state.end_hour))
    st.session_state.lunch_start = st.selectbox("Inicio almuerzo", HOURS, index=idx_of(st.session_state.lunch_start))
    st.session_state.lunch_end = st.selectbox("Fin almuerzo", HOURS, index=idx_of(st.session_state.lunch_end))

    st.divider()
    st.subheader("Save / Load Config (JSON)")
    if st.button("Save Config (JSON)"):
        data = {
            "clients": st.session_state.clients,
            "buyers": st.session_state.buyers,
            "appointments": st.session_state.appointments,
            "start_hour": st.session_state.start_hour,
            "end_hour": st.session_state.end_hour,
            "lunch_start": st.session_state.lunch_start,
            "lunch_end": st.session_state.lunch_end,
            "selected_days": st.session_state.selected_days,
            "time_windows": st.session_state.time_windows,
        }
        json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
        st.download_button("Download config JSON", data=json_bytes, file_name="ubagofish_config.json", mime="application/json")

    uploaded = st.file_uploader("Load Config (JSON)", type=["json"])
    if uploaded is not None:
        try:
            loaded = json.load(uploaded)
            st.session_state.clients = loaded.get("clients", st.session_state.clients)
            st.session_state.buyers = loaded.get("buyers", st.session_state.buyers)
            st.session_state.appointments = loaded.get("appointments", st.session_state.appointments)
            st.session_state.start_hour = loaded.get("start_hour", st.session_state.start_hour)
            st.session_state.end_hour = loaded.get("end_hour", st.session_state.end_hour)
            st.session_state.lunch_start = loaded.get("lunch_start", st.session_state.lunch_start)
            st.session_state.lunch_end = loaded.get("lunch_end", st.session_state.lunch_end)
            st.session_state.selected_days = loaded.get("selected_days", st.session_state.selected_days)
            st.session_state.time_windows = loaded.get("time_windows", st.session_state.time_windows)
            autosave(); st.success("Configuración cargada desde JSON.")
        except Exception as e:
            st.error(f"Error cargando JSON: {e}")

    st.divider()
    with st.expander("🗑️ Editar / Borrar Citas"):
        if st.button("Borrar TODAS las citas"):
            st.session_state.appointments.clear(); autosave(); st.warning("Todas las citas fueron eliminadas.")
        buyer_clear = st.selectbox("Borrar citas de Buyer", [""] + st.session_state.buyers, key="clear_buyer")
        if st.button("Borrar citas del Buyer seleccionado") and buyer_clear:
            st.session_state.appointments = [a for a in st.session_state.appointments if a["buyer"] != buyer_clear]
            autosave(); st.warning(f"Citas de {buyer_clear} eliminadas.")
        client_clear = st.selectbox("Borrar citas de Client", [""] + st.session_state.clients, key="clear_client")
        if st.button("Borrar citas del Client seleccionado") and client_clear:
            st.session_state.appointments = [a for a in st.session_state.appointments if a["client"] != client_clear]
            autosave(); st.warning(f"Citas de {client_clear} eliminadas.")

# -------------------------
# Tabs (Randomize / Manual)
# -------------------------

tab_random, tab_manual = st.tabs(["🎲 Generador Aleatorio", "✏️ Agendar Manualmente"]) 

# -------------------------
# Randomizer with heuristics
# -------------------------
with tab_random:
    st.subheader("🎲 Generar citas aleatorias con descansos y balance por días")
    selected_buyers = []
    col1, col2 = st.columns([1,1])
    with col1:
        if "buyers_random" not in st.session_state:
            st.session_state.buyers_random = [""]
        for i,_ in enumerate(st.session_state.buyers_random):
            buyer = st.selectbox(f"Buyer {i+1}", st.session_state.buyers, key=f"buyer_random_{i}")
            if buyer:
                selected_buyers.append(buyer)
        if st.button("Agregar otro Buyer"):
            st.session_state.buyers_random.append("")
    with col2:
        selected_clients = st.multiselect("Seleccionar Clients", st.session_state.clients)

    st.markdown("### Ventanas Horarias por Buyer (opcional)")
    for buyer in selected_buyers:
        st.markdown(f"**{buyer}**")
        st.session_state.time_windows.setdefault(buyer, {})
        for day in st.session_state.selected_days:
            col_from, col_to = st.columns(2)
            with col_from:
                start = st.selectbox(
                    f"{day} desde",
                    HOURS,
                    key=f"{buyer}_{day}_start",
                    index=idx_of(st.session_state.time_windows.get(buyer, {}).get(day, {}).get("start", st.session_state.start_hour)),
                )
            with col_to:
                end = st.selectbox(
                    f"{day} hasta",
                    HOURS,
                    key=f"{buyer}_{day}_end",
                    index=idx_of(st.session_state.time_windows.get(buyer, {}).get(day, {}).get("end", st.session_state.end_hour)),
                )
            st.session_state.time_windows[buyer][day] = {"start": start, "end": end}
    autosave()

    st.divider()
    colA, colB, colC = st.columns([1,1,1])
    with colA:
        interval = st.selectbox("Duración de cita (min)", [30, 60], key="interval")
    with colB:
        appts_before_rest = st.number_input("Citas antes de descanso", min_value=1, max_value=6, value=2, step=1, key="appts_before_rest")
    with colC:
        rest_slots = st.number_input("Duración del descanso (slots)", min_value=1, max_value=3, value=1, step=1, key="rest_slots")

    def gen_slots_for(buyer: str, day: str) -> list[str]:
        """Return ordered list of time strings honoring buyer's day window, global window, lunch and interval."""
        start = st.session_state.time_windows.get(buyer, {}).get(day, {}).get("start", st.session_state.start_hour)
        end = st.session_state.time_windows.get(buyer, {}).get(day, {}).get("end", st.session_state.end_hour)
        start_idx = max(idx_of(start), idx_of(st.session_state.start_hour))
        end_idx = min(idx_of(end), idx_of(st.session_state.end_hour))
        step = interval // 30
        slots = []
        for i in range(start_idx, end_idx):
            if i % step != 0:
                continue
            t = HOURS[i]
            if not is_in_lunch_break(t):
                slots.append(t)
        return slots

    def remove_unlocked_appointments_for(buyers: list[str]):
        """Remove only unlocked (non-locked) appointments that involve selected buyers across all days."""
        st.session_state.appointments = [a for a in st.session_state.appointments if (a["buyer"] not in buyers) or a.get("locked", False)]

    def balanced_bucket(count: int, days: list[str]) -> dict:
        """Return {day: n} appts distributed as evenly as possible."""
        if not days:
            return {}
        q, r = divmod(count, len(days))
        alloc = {d: q for d in days}
        for i in range(r):
            alloc[days[i % len(days)]] += 1
        return alloc

    def place_for_day(buyer: str, day: str, clients_for_day: list[str]):
        """Place clients on that day respecting rest cadence and existing locked blocks."""
        # Map of occupied times for that day (including locked and those placed earlier in this run)
        taken = {a["time"]: a for a in st.session_state.appointments if a["day"] == day}
        slots = gen_slots_for(buyer, day)
        placed = 0
        cadence_count = 0
        i = 0
        ci = 0
        while i < len(slots) and ci < len(clients_for_day):
            t = slots[i]
            # Enforce cadence: after appts_before_rest, skip rest_slots slots
            if cadence_count >= appts_before_rest:
                i += rest_slots
                cadence_count = 0
                continue

            # Slot availability: slot must be free and no conflicts for client/buyer and not occupied by locked
            if t not in taken and is_slot_free(clients_for_day[ci], buyer, day, t):
                st.session_state.appointments.append({
                    "client": clients_for_day[ci],
                    "buyer": buyer,
                    "day": day,
                    "time": t,
                    "locked": False,
                })
                taken[t] = True
                placed += 1
                cadence_count += 1
                ci += 1
                i += 1
            else:
                i += 1
        return placed

    if st.button("Generar citas aleatorias"):
        # remove previous unlocked appointments for the selected buyers so we can reflow
        remove_unlocked_appointments_for(selected_buyers)
        days_pool = st.session_state.selected_days[:]
        for buyer in selected_buyers:
            total = len(selected_clients)
            if total == 0:
                continue
            alloc = balanced_bucket(total, days_pool)
            # Distribute specific clients into day buckets in round-robin fashion
            day_lists = {d: [] for d in days_pool}
            day_cycle = [d for d, n in alloc.items() for _ in range(n)]
            for idx, client in enumerate(selected_clients):
                if idx < len(day_cycle):
                    day_lists[day_cycle[idx]].append(client)
                else:
                    # fallback: append to first day
                    day_lists[days_pool[0]].append(client)
            for d in days_pool:
                if day_lists[d]:
                    place_for_day(buyer, d, day_lists[d])
        autosave(); st.success("Citas generadas y reacomodadas (locked respetadas, días balanceados, descansos aplicados).")

# -------------------------
# Manual scheduling (locked)
# -------------------------
with tab_manual:
    st.subheader("✏️ Agendar Manualmente (bloquea el horario)")
    if not st.session_state.buyers or not st.session_state.clients:
        st.info("Añade buyers y clients en la barra lateral antes de crear citas manuales.")
    else:
        buyer_manual = st.selectbox("Buyer", st.session_state.buyers, key="buyer_manual")
        client_manual = st.selectbox("Client", st.session_state.clients, key="client_manual")
        dia_manual = st.selectbox("Día", st.session_state.selected_days, key="dia_manual")
        valid_times = [h for h in HOURS if idx_of(st.session_state.start_hour) <= idx_of(h) < idx_of(st.session_state.end_hour)]
        hora_manual = st.selectbox("Hora", valid_times, key="hora_manual")
        if st.button("Agendar cita manual"):
            if is_in_lunch_break(hora_manual):
                st.warning("No se pueden agendar durante el almuerzo.")
            elif not is_slot_free(client_manual, buyer_manual, dia_manual, hora_manual):
                st.warning("El Buyer o Client ya tiene cita a esa hora.")
            else:
                appt = {"client": client_manual, "buyer": buyer_manual, "day": dia_manual, "time": hora_manual, "locked": True}
                exists = any(a for a in st.session_state.appointments if a["client"]==client_manual and a["buyer"]==buyer_manual and a["day"]==dia_manual and a["time"]==hora_manual)
                if exists:
                    st.warning("Esta cita ya está agendada.")
                else:
                    st.session_state.appointments.append(appt); autosave(); st.success("Cita manual agendada y bloqueada.")

# -------------------------
# Calendar view
# -------------------------
st.subheader("📅 Calendario de Citas")
if st.session_state.appointments:
    data = []
    for day in DAYS:
        row = {"Hora": day}
        appts_day = [a for a in st.session_state.appointments if a["day"] == day]
        for time in HOURS[idx_of(st.session_state.start_hour):idx_of(st.session_state.end_hour)]:
            if is_in_lunch_break(time):
                row[time] = "LUNCH BREAK"
            else:
                labels = []
                for a in appts_day:
                    if a["time"] == time:
                        label = f"{a['buyer']} - {a['client']}" + (" 🔒" if a.get("locked") else "")
                        labels.append(label)
                row[time] = "; ".join(labels)
        data.append(row)
    df = pd.DataFrame(data).set_index("Hora").T
    st.dataframe(df, use_container_width=True)
else:
    st.info("No hay citas programadas aún.")

# -------------------------
# Export to Excel (two sheets per day: ByBuyer, ByClient)
# -------------------------
if st.button("📤 Export Schedule (Excel)"):
    times = HOURS[idx_of(st.session_state.start_hour):idx_of(st.session_state.end_hour)]
    buyers = st.session_state.buyers[:]
    clients = st.session_state.clients[:]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for day in st.session_state.selected_days:
            # buyer view
            df_b = pd.DataFrame(index=times, columns=buyers)
            for t in times:
                for b in buyers:
                    df_b.at[t, b] = "LUNCH BREAK" if is_in_lunch_break(t) else ""
            # client view
            df_c = pd.DataFrame(index=times, columns=clients)
            for t in times:
                for c in clients:
                    df_c.at[t, c] = "LUNCH BREAK" if is_in_lunch_break(t) else ""
            # fill day's appointments
            for a in [x for x in st.session_state.appointments if x["day"] == day]:
                t = a["time"]
                b = a["buyer"]
                c = a["client"]
                client_marker = f"{c}{'*' if a.get('locked') else ''}"
                buyer_marker = f"{b}{'*' if a.get('locked') else ''}"
                if b in buyers:
                    df_b.at[t, b] = client_marker
                if c in clients:
                    df_c.at[t, c] = buyer_marker

            df_b_reset = df_b.reset_index().rename(columns={"index": "Time"})
            df_c_reset = df_c.reset_index().rename(columns={"index": "Time"})
            df_b_reset.to_excel(writer, sheet_name=f"ByBuyer_{day}", index=False)
            df_c_reset.to_excel(writer, sheet_name=f"ByClient_{day}", index=False)

        # summary sheets
        df_all = pd.DataFrame(st.session_state.appointments)
        if not df_all.empty:
            df_all = df_all.rename(columns={"client": "Client", "buyer": "Buyer", "day": "Día", "time": "Hora", "locked": "Locked"})
            df_all["Count"] = 1
            df_all.groupby("Client")["Count"].sum().reset_index().to_excel(writer, sheet_name="Summary_Clients", index=False)
            df_all.groupby("Buyer")["Count"].sum().reset_index().to_excel(writer, sheet_name="Summary_Buyers", index=False)

    # style the workbook (headers & lunch grey)
    output.seek(0)
    wb = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    lunch_fill = PatternFill("solid", fgColor="D9D9D9")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if cell.value == "LUNCH BREAK":
                    cell.fill = lunch_fill
    final = BytesIO(); wb.save(final); final.seek(0)
    st.download_button("Download Schedule Excel", data=final, file_name="UbagoFish_Schedule_v2.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -------------------------
# In-place editor for existing appts
# -------------------------
with st.expander("🔧 Editar Citas", expanded=st.session_state.edit_expander_open):
    st.session_state.edit_expander_open = True
    if st.session_state.appointments:
        options = [
            f"{a['client']} con {a['buyer']} ({a['day']} {a['time']})" + (" [locked]" if a.get("locked") else "")
            for a in st.session_state.appointments
        ]
        sel = st.selectbox("Seleccionar cita para editar", options)
        if sel:
            idx = options.index(sel)
            a = st.session_state.appointments[idx]
            new_b = st.selectbox("Nuevo Buyer", st.session_state.buyers, index=st.session_state.buyers.index(a["buyer"]))
            new_c = st.selectbox("Nuevo Client", st.session_state.clients, index=st.session_state.clients.index(a["client"]))
            new_d = st.selectbox("Nuevo Día", st.session_state.selected_days, index=st.session_state.selected_days.index(a["day"]))
            new_h = st.selectbox("Nueva Hora", HOURS, index=HOURS.index(a["time"]))
            new_locked = st.checkbox("Marcar como locked (bloqueado)", value=a.get("locked", False))
            if st.button("Guardar cambios"):
                if is_in_lunch_break(new_h):
                    st.warning("No se pueden agendar durante el almuerzo.")
                elif not (a["day"] == new_d and a["time"] == new_h and a["buyer"] == new_b and a["client"] == new_c) and not is_slot_free(new_c, new_b, new_d, new_h):
                    st.warning("El Buyer o Client ya tiene cita a esa hora.")
                else:
                    st.session_state.appointments[idx] = {"client": new_c, "buyer": new_b, "day": new_d, "time": new_h, "locked": new_locked}
                    autosave(); st.success("Cita editada.")
    else:
        st.info("No hay citas para editar.")
