
import streamlit as st
import pandas as pd
import json
import os
from io import BytesIO
from random import choice
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="UbagoFish Scheduler", layout="wide")
st.title("🐟 UbagoFish Scheduler")
st.caption("Version 1.2 – Full App with Styled Excel Export")

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
HOURS = [f"{h:02d}:{m:02d}" for h in range(6, 22) for m in (0, 30)]
DATA_FILE = "ubagofish_data.json"

# Initialize session state
for key in ["proveedores", "empresas", "appointments"]:
    if key not in st.session_state:
        st.session_state[key] = []
if "edit_mode" not in st.session_state:
    st.session_state.edit_mode = False
if "appointment_to_edit" not in st.session_state:
    st.session_state.appointment_to_edit = None
if "start_hour" not in st.session_state:
    st.session_state.start_hour = "06:00"
if "end_hour" not in st.session_state:
    st.session_state.end_hour = "21:30"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            data = json.load(f)
            st.session_state.proveedores = data.get("proveedores", [])
            st.session_state.empresas = data.get("empresas", [])
            st.session_state.appointments = [tuple(app) for app in data.get("appointments", [])]

def save_data():
    with open(DATA_FILE, "w") as f:
        json.dump({
            "proveedores": st.session_state.proveedores,
            "empresas": st.session_state.empresas,
            "appointments": st.session_state.appointments,
        }, f)

def autosave():
    save_data()

load_data()

# Sidebar: Empresas/Proveedores
st.sidebar.header("Empresas y Proveedores")
empresas_input = st.sidebar.text_area("Empresas (una por línea)", "\n".join(st.session_state.empresas))
st.session_state.empresas = [e.strip() for e in empresas_input.splitlines() if e.strip()]
proveedores_input = st.sidebar.text_area("Proveedores (uno por línea)", "\n".join(st.session_state.proveedores))
st.session_state.proveedores = [p.strip() for p in proveedores_input.splitlines() if p.strip()]

if st.sidebar.button("Guardar nombres"):
    autosave()
    st.sidebar.success("Empresas y Proveedores guardados.")
if st.sidebar.button("Guardar progreso manualmente"):
    save_data()
    st.sidebar.success("Progreso guardado.")

# Lunch break
lunch_start = st.sidebar.selectbox("Inicio del almuerzo (Bloqueo)", HOURS, index=12)
lunch_end = st.sidebar.selectbox("Fin del almuerzo (Bloqueo)", HOURS, index=14)
lunch_start_idx, lunch_end_idx = HOURS.index(lunch_start), HOURS.index(lunch_end)
def is_in_lunch_break(time_val): return lunch_start_idx <= HOURS.index(time_val) < lunch_end_idx
st.session_state.appointments = [appt for appt in st.session_state.appointments if not is_in_lunch_break(appt[3])]

# Manage appointments
st.sidebar.subheader("Administrar citas")
action = st.sidebar.selectbox("Acción", ["Ninguna", "Limpiar todo", "Limpiar por Empresa", "Limpiar por Proveedor", "Editar cita"])
if action == "Limpiar todo" and st.sidebar.button("Ejecutar"):
    st.session_state.appointments = []
    autosave()
    st.sidebar.success("Todas las citas eliminadas.")
elif action == "Limpiar por Empresa":
    emp = st.sidebar.selectbox("Selecciona Empresa", st.session_state.empresas)
    if st.sidebar.button("Limpiar citas de esta Empresa"):
        st.session_state.appointments = [appt for appt in st.session_state.appointments if appt[1] != emp]
        autosave()
        st.sidebar.success(f"Citas de {emp} eliminadas.")
elif action == "Limpiar por Proveedor":
    prov = st.sidebar.selectbox("Selecciona Proveedor", st.session_state.proveedores)
    if st.sidebar.button("Limpiar citas de este Proveedor"):
        st.session_state.appointments = [appt for appt in st.session_state.appointments if appt[0] != prov]
        autosave()
        st.sidebar.success(f"Citas de {prov} eliminadas.")
elif action == "Editar cita" and st.session_state.appointments:
    appt_list = [f"{p} - {e} ({d} {t})" for p,e,d,t in st.session_state.appointments]
    appt_choice = st.sidebar.selectbox("Selecciona cita", appt_list)
    if st.sidebar.button("Editar esta cita"):
        idx = appt_list.index(appt_choice)
        st.session_state.edit_mode = True
        st.session_state.appointment_to_edit = idx

if st.session_state.edit_mode and st.session_state.appointment_to_edit is not None:
    st.subheader("Editar cita")
    old_p, old_e, old_d, old_t = st.session_state.appointments[st.session_state.appointment_to_edit]
    new_p = st.selectbox("Proveedor", st.session_state.proveedores, index=st.session_state.proveedores.index(old_p))
    new_e = st.selectbox("Empresa", st.session_state.empresas, index=st.session_state.empresas.index(old_e))
    new_d = st.selectbox("Día", DAYS, index=DAYS.index(old_d))
    new_t = st.selectbox("Hora", HOURS, index=HOURS.index(old_t))
    if st.button("Guardar cambios"):
        if is_in_lunch_break(new_t):
            st.warning("No se pueden agendar durante almuerzo.")
        else:
            st.session_state.appointments[st.session_state.appointment_to_edit] = (new_p,new_e,new_d,new_t)
            st.session_state.edit_mode = False
            st.session_state.appointment_to_edit = None
            autosave()
            st.success("Cita editada correctamente.")

# Tabs for scheduling
tab1, tab2 = st.tabs(["🎲 Generador Aleatorio", "📝 Agendar Manualmente"])
with tab1:
    st.subheader("🎲 Generar citas aleatorias")
    selected_empresas = []
    col1, col2 = st.columns([1, 1])
    with col1:
        if "empresas_random" not in st.session_state:
            st.session_state.empresas_random = [""]
        for i, _ in enumerate(st.session_state.empresas_random):
            empresa = st.selectbox(f"Empresa {i+1}", options=st.session_state.empresas, key=f"empresa_random_{i}")
            selected_empresas.append(empresa)
        if st.button("➕ Agregar otra empresa"):
            st.session_state.empresas_random.append("")
    with col2:
        selected_proveedores = st.multiselect("Seleccionar proveedores", options=st.session_state.proveedores)
    col3, col4 = st.columns(2)
    with col3:
        day_range = st.multiselect("Seleccionar días", DAYS, default=["Monday", "Tuesday", "Wednesday"])
    with col4:
        st.session_state.start_hour = st.selectbox("Inicio del día", HOURS, index=6)
        st.session_state.end_hour = st.selectbox("Fin del día", HOURS, index=20)
    interval = st.selectbox("Duración de la cita (min)", [30, 60])
    if st.button("🔀 Generar citas aleatorias"):
        start_idx, end_idx = HOURS.index(st.session_state.start_hour), HOURS.index(st.session_state.end_hour)
        available_slots = [h for h in HOURS[start_idx:end_idx] if not is_in_lunch_break(h)]
        available_slots = [h for h in available_slots if HOURS.index(h) % (interval // 30) == 0]
        for empresa in selected_empresas:
            for proveedor in selected_proveedores:
                for day in day_range:
                    attempts = 0
                    while attempts < 5 and available_slots:
                        time = choice(available_slots)
                        if (proveedor, empresa, day, time) not in st.session_state.appointments:
                            st.session_state.appointments.append((proveedor, empresa, day, time))
                            break
                        attempts += 1
        autosave()
        st.success("Citas generadas y guardadas.")

with tab2:
    st.subheader("📝 Agendar manualmente")
    empresa_manual = st.selectbox("Empresa", st.session_state.empresas, key="empresa_manual")
    proveedor_manual = st.selectbox("Proveedor", st.session_state.proveedores, key="proveedor_manual")
    dia_manual = st.selectbox("Día", DAYS, key="dia_manual")
    hora_manual = st.selectbox("Hora", HOURS, key="hora_manual")
    if st.button("➕ Agendar cita manual"):
        if is_in_lunch_break(hora_manual):
            st.warning("No se pueden agendar durante el almuerzo.")
        else:
            new_appointment = (proveedor_manual, empresa_manual, dia_manual, hora_manual)
            if new_appointment in st.session_state.appointments:
                st.warning("Esta cita ya está agendada.")
            else:
                st.session_state.appointments.append(new_appointment)
                autosave()
                st.success("Cita agendada exitosamente.")

# Weekly calendar
st.markdown("---")
st.subheader("📅 Calendario de citas (semanal)")
if st.session_state.appointments:
    data = []
    start_idx, end_idx = HOURS.index(st.session_state.start_hour), HOURS.index(st.session_state.end_hour)
    visible_hours = HOURS[start_idx:end_idx]
    for day in DAYS:
        row = {"Hora": day}
        appts = [(p, e, d, t) for (p, e, d, t) in st.session_state.appointments if d == day]
        for time in visible_hours:
            if is_in_lunch_break(time):
                row[time] = "LUNCH BREAK"
            else:
                cell = [f"{e} - {p}" for (p, e, d, t) in appts if t == time]
                row[time] = "; ".join(cell)
        data.append(row)
    df_week = pd.DataFrame(data).set_index("Hora").T
    st.dataframe(df_week.style.apply(lambda col: ["background-color: #d9d9d9" if v == "LUNCH BREAK" else "" for v in col], axis=0), use_container_width=True)
else:
    st.info("No hay citas programadas aún.")

# Styled Excel export
if st.button("📤 Exportar Horario a Excel"):
    def export_schedule_excel():
        appointments_df = pd.DataFrame(st.session_state.appointments, columns=["Proveedor", "Empresa", "Día", "Hora"])
        output = BytesIO()

        def style_sheet(ws):
            header_fill = PatternFill("solid", fgColor="305496")
            header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
            lunch_fill = PatternFill("solid", fgColor="D9D9D9")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
                for cell in row:
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=11)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == "LUNCH BREAK":
                        cell.fill = lunch_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        def write_schedule(writer, sheet_prefix, columns_key, group_key):
            for day in appointments_df["Día"].unique():
                day_df = appointments_df[appointments_df["Día"] == day]
                start_idx, end_idx = HOURS.index(st.session_state.start_hour), HOURS.index(st.session_state.end_hour)
                timeframes = HOURS[start_idx:end_idx]
                result = pd.DataFrame({"Time": timeframes})
                for item in st.session_state[columns_key]:
                    df_item = day_df[day_df[group_key] == item]
                    paired_list = []
                    for t in timeframes:
                        if is_in_lunch_break(t):
                            paired_list.append("LUNCH BREAK")
                        else:
                            row = df_item[df_item["Hora"] == t]
                            paired_list.append(", ".join(row["Empresa" if group_key=="Proveedor" else "Proveedor"].tolist()) if not row.empty else "")
                    result[item] = paired_list
                result.to_excel(writer, sheet_name=f"{sheet_prefix}_{day}", index=False)

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            write_schedule(writer, "Proveedores", "proveedores", "Proveedor")
            write_schedule(writer, "Empresas", "empresas", "Empresa")
            writer.save()
        # Apply styles
        output.seek(0)
        wb = load_workbook(output)
        for ws in wb.worksheets:
            style_sheet(ws)
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        st.download_button("Descargar Horario Completo", data=final_output, file_name="UbagoFish_Schedule_Styled.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    export_schedule_excel()
