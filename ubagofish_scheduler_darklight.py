
import streamlit as st
import pandas as pd
import json, os
from io import BytesIO
from random import choice
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

st.set_page_config(page_title="UbagoFish Scheduler - Dark/Light v1.4", layout="wide")

# Initialize session state for dark mode
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False

# Sidebar toggle
with st.sidebar:
    toggle = st.checkbox("🌙 Activar modo oscuro", value=st.session_state.dark_mode)
    if toggle != st.session_state.dark_mode:
        st.session_state.dark_mode = toggle
        st.experimental_rerun()

# Dark/Light palettes
if st.session_state.dark_mode:
    palette = {
        "bg_color": "#1E293B",
        "text_color": "#FFFFFF",
        "header_color": "#F8FAFC",
        "card_color": "#334155",
        "alt_row_color": "#334155",
        "hover_color": "#475569",
        "sidebar_color": "#1E293B"
    }
else:
    palette = {
        "bg_color": "#F8FAFC",
        "text_color": "#1E293B",
        "header_color": "#1E293B",
        "card_color": "#FFFFFF",
        "alt_row_color": "#F2F6FA",
        "hover_color": "#E6FFF5",
        "sidebar_color": "#F1F5F9"
    }

# Load CSS
with open("style.css") as f:
    css_template = f.read()
css = css_template.format(**palette)
st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)

# Title
st.title("🐟 UbagoFish Scheduler – Dark/Light v1.4")
st.caption(f"Current time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

# Constants and state
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
HOURS = [f"{h:02d}:{m:02d}" for h in range(6, 22) for m in (0,30)]
DATA_FILE = "ubagofish_data.json"

for key in ["clients", "buyers", "appointments"]:
    if key not in st.session_state: st.session_state[key] = []
if "edit_expander_open" not in st.session_state: st.session_state.edit_expander_open = False
if "start_hour" not in st.session_state: st.session_state.start_hour = "06:00"
if "end_hour" not in st.session_state: st.session_state.end_hour = "21:30"
if "lunch_start" not in st.session_state: st.session_state.lunch_start = "12:00"
if "lunch_end" not in st.session_state: st.session_state.lunch_end = "14:00"
if "selected_days" not in st.session_state: st.session_state.selected_days = ["Monday", "Tuesday"]
if "time_windows" not in st.session_state: st.session_state.time_windows = {}

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            data = json.load(f)
            st.session_state.clients = data.get("clients", [])
            st.session_state.buyers = data.get("buyers", [])
            st.session_state.appointments = [tuple(app) for app in data.get("appointments", [])]
            st.session_state.lunch_start = data.get("lunch_start", "12:00")
            st.session_state.lunch_end = data.get("lunch_end", "14:00")
            st.session_state.selected_days = data.get("selected_days", ["Monday", "Tuesday"])
            st.session_state.time_windows = data.get("time_windows", {})
def save_data():
    with open(DATA_FILE, "w") as f:
        json.dump({"clients": st.session_state.clients, "buyers": st.session_state.buyers,
                   "appointments": st.session_state.appointments,
                   "lunch_start": st.session_state.lunch_start,
                   "lunch_end": st.session_state.lunch_end,
                   "selected_days": st.session_state.selected_days,
                   "time_windows": st.session_state.time_windows}, f)
def autosave(): save_data()
load_data()

# Utility
lunch_start_idx, lunch_end_idx = HOURS.index(st.session_state.lunch_start), HOURS.index(st.session_state.lunch_end)
def is_in_lunch_break(t): return lunch_start_idx <= HOURS.index(t) < lunch_end_idx
st.session_state.appointments = [a for a in st.session_state.appointments if not is_in_lunch_break(a[3])]

# Sidebar inputs
with st.sidebar:
    st.header("Buyers & Clients")
    buyers_input = st.text_area("Buyers (uno por línea)", "\n".join(st.session_state.buyers))
    st.session_state.buyers = [b.strip() for b in buyers_input.splitlines() if b.strip()]
    clients_input = st.text_area("Clients (uno por línea)", "\n".join(st.session_state.clients))
    st.session_state.clients = [c.strip() for c in clients_input.splitlines() if c.strip()]
    if st.button("Guardar nombres"): autosave(); st.success("Datos guardados.")
    st.subheader("Configuración")
    st.session_state.lunch_start = st.selectbox("Inicio almuerzo", HOURS, index=HOURS.index(st.session_state.lunch_start))
    st.session_state.lunch_end = st.selectbox("Fin almuerzo", HOURS, index=HOURS.index(st.session_state.lunch_end))
    st.session_state.selected_days = st.multiselect("Días a programar", DAYS, default=st.session_state.selected_days)

# Scheduler Tabs
tab_random, tab_manual = st.tabs(["🎲 Generador Aleatorio", "✏️ Agendar Manualmente"])

with tab_random:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("🎲 Generador Aleatorio")
    selected_buyers = []
    col1, col2 = st.columns([1,1])
    with col1:
        if "buyers_random" not in st.session_state: st.session_state.buyers_random = [""]
        for i,_ in enumerate(st.session_state.buyers_random):
            buyer = st.selectbox(f"Buyer {i+1}", st.session_state.buyers, key=f"buyer_random_{i}")
            selected_buyers.append(buyer)
        if st.button("Agregar otro Buyer"): st.session_state.buyers_random.append("")
    with col2:
        selected_clients = st.multiselect("Seleccionar Clients", st.session_state.clients)
    st.markdown("### Ventanas Horarias (opcional)")
    for buyer in selected_buyers:
        st.markdown(f"**{buyer}**")
        st.session_state.time_windows.setdefault(buyer, {})
        for day in st.session_state.selected_days:
            col_from, col_to = st.columns(2)
            with col_from:
                start = st.selectbox(f"{day} desde", HOURS, key=f"{buyer}_{day}_start",
                                     index=HOURS.index(st.session_state.time_windows.get(buyer, {}).get(day, {}).get("start", st.session_state.start_hour)))
            with col_to:
                end = st.selectbox(f"{day} hasta", HOURS, key=f"{buyer}_{day}_end",
                                   index=HOURS.index(st.session_state.time_windows.get(buyer, {}).get(day, {}).get("end", st.session_state.end_hour)))
            st.session_state.time_windows[buyer][day] = {"start": start, "end": end}
    autosave()
    interval = st.selectbox("Duración de cita (min)", [30, 60])
    if st.button("Generar citas aleatorias"):
        with st.spinner("Generando citas..."):
            for buyer in selected_buyers:
                for client in selected_clients:
                    for day in st.session_state.selected_days:
                        start = st.session_state.time_windows.get(buyer, {}).get(day, {}).get("start", st.session_state.start_hour)
                        end = st.session_state.time_windows.get(buyer, {}).get(day, {}).get("end", st.session_state.end_hour)
                        slots = [h for h in HOURS[HOURS.index(start):HOURS.index(end)] if not is_in_lunch_break(h) and HOURS.index(h) % (interval//30)==0]
                        attempts = 0
                        while attempts < 5 and slots:
                            t = choice(slots)
                            if (client,buyer,day,t) not in st.session_state.appointments:
                                st.session_state.appointments.append((client,buyer,day,t)); break
                            attempts += 1
            autosave(); st.success("Citas generadas.")
    st.markdown('</div>', unsafe_allow_html=True)

with tab_manual:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("✏️ Agendar Manualmente")
    buyer_manual = st.selectbox("Buyer", st.session_state.buyers, key="buyer_manual")
    client_manual = st.selectbox("Client", st.session_state.clients, key="client_manual")
    dia_manual = st.selectbox("Día", DAYS, key="dia_manual")
    hora_manual = st.selectbox("Hora", HOURS, key="hora_manual")
    if st.button("Agendar cita manual"):
        if is_in_lunch_break(hora_manual): st.warning("No se pueden agendar durante el almuerzo.")
        else:
            appt = (client_manual,buyer_manual,dia_manual,hora_manual)
            if appt in st.session_state.appointments: st.warning("Esta cita ya está agendada.")
            else:
                st.session_state.appointments.append(appt); autosave(); st.success("Cita agendada exitosamente.")
    st.markdown('</div>', unsafe_allow_html=True)

# Calendar
st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("📅 Calendario de Citas")
if st.session_state.appointments:
    data=[]
    for day in DAYS:
        row={"Hora":day}; appts=[a for a in st.session_state.appointments if a[2]==day]
        for time in HOURS[HOURS.index(st.session_state.start_hour):HOURS.index(st.session_state.end_hour)]:
            if is_in_lunch_break(time): row[time]="LUNCH BREAK"
            else:
                row[time]="; ".join([f"<b style='color:#38BDF8'>{b}</b> - <span style='color:#F87171'>{c}</span>" for c,b,d,t in appts if t==time])
        data.append(row)
    df=pd.DataFrame(data).set_index("Hora").T
    st.write(df.to_html(escape=False), unsafe_allow_html=True)
else: st.info("No hay citas programadas aún.")
st.markdown('</div>', unsafe_allow_html=True)

# Floating actions
col_act1, col_act2 = st.columns(2)
with col_act1:
    if st.button("Limpiar TODO"):
        st.session_state.appointments = []; autosave(); st.success("Todas las citas eliminadas.")
with col_act2:
    if st.button("📤 Exportar a Excel"):
        with st.spinner("Exportando horario..."):
            df_all = pd.DataFrame(st.session_state.appointments, columns=["Client","Buyer","Día","Hora"])
            output = BytesIO()
            def style_ws(ws):
                header_fill=PatternFill("solid",fgColor="305496"); header_font=Font(color="FFFFFF",bold=True,name="Calibri",size=11)
                lunch_fill=PatternFill("solid",fgColor="D9D9D9"); border=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
                for r,row in enumerate(ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=ws.max_column),start=1):
                    for cell in row:
                        cell.border=border; cell.font=Font(name="Calibri",size=11)
                        if r==1: cell.fill=header_fill; cell.font=header_font; cell.alignment=Alignment(horizontal="center",vertical="center")
                        if cell.value=="LUNCH BREAK": cell.fill=lunch_fill
                        cell.alignment=Alignment(horizontal="center",vertical="center")
                for col in ws.columns:
                    max_len=max(len(str(c.value)) if c.value else 0 for c in col)
                    ws.column_dimensions[col[0].column_letter].width=max_len+2
            def write_sheet(writer,prefix,key,group):
                for day in df_all["Día"].unique():
                    df_day=df_all[df_all["Día"]==day]; times=HOURS
                    result=pd.DataFrame({"Time":times})
                    for item in st.session_state[key]:
                        cells=[]; sub=df_day[df_day[group]==item]
                        for t in times:
                            if is_in_lunch_break(t): cells.append("LUNCH BREAK")
                            else:
                                sub_row=sub[sub["Hora"]==t]; cells.append(", ".join(sub_row["Buyer" if group=="Client" else "Client"].tolist()) if not sub_row.empty else "")
                        result[item]=cells
                    result.to_excel(writer,sheet_name=f"{prefix}_{day}",index=False)
            def write_summary(writer,prefix,column):
                counts=df_all[column].value_counts().reset_index()
                counts.columns=[column,"Total Citas"]
                counts.to_excel(writer,sheet_name=f"Summary_{prefix}",index=False)
            with pd.ExcelWriter(output,engine="openpyxl") as writer:
                write_sheet(writer,"Clients","clients","Client")
                write_sheet(writer,"Buyers","buyers","Buyer")
                write_summary(writer,"Clients","Client")
                write_summary(writer,"Buyers","Buyer")
            output.seek(0); wb=load_workbook(output)
            for ws in wb.worksheets: style_ws(ws)
            final=BytesIO(); wb.save(final); final.seek(0)
            st.download_button("Descargar Horario Completo", data=final,
                               file_name="UbagoFish_Schedule_DarkLight_v14.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Editar citas
with st.expander("🔧 Editar Citas", expanded=st.session_state.edit_expander_open):
    st.session_state.edit_expander_open = True
    if st.session_state.appointments:
        appt_options = [f"{c} con {b} ({d} a las {h})" for c,b,d,h in st.session_state.appointments]
        selected_edit = st.selectbox("Selecciona una cita para editar", appt_options)
        if selected_edit:
            idx = appt_options.index(selected_edit)
            c,b,d,h = st.session_state.appointments[idx]
            buyer_idx = st.session_state.buyers.index(b) if b in st.session_state.buyers else 0
            client_idx = st.session_state.clients.index(c) if c in st.session_state.clients else 0
            new_buyer = st.selectbox("Nuevo Buyer", st.session_state.buyers, index=buyer_idx)
            new_client = st.selectbox("Nuevo Client", st.session_state.clients, index=client_idx)
            new_day = st.selectbox("Nuevo Día", DAYS, index=DAYS.index(d))
            new_time = st.selectbox("Nueva Hora", HOURS, index=HOURS.index(h))
            if st.button("Guardar cambios"):
                if is_in_lunch_break(new_time): st.warning("No se pueden agendar durante el almuerzo.")
                else:
                    new_appt = (new_client,new_buyer,new_day,new_time)
                    if new_appt in st.session_state.appointments and new_appt != st.session_state.appointments[idx]:
                        st.warning("Ya existe una cita con estos detalles.")
                    else:
                        st.session_state.appointments[idx] = new_appt; autosave(); st.success("Cita editada exitosamente.")
